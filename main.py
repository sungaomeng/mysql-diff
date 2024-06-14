import pymysql
import requests
import json
import csv
import os
import logging
from requests_toolbelt.multipart.encoder import MultipartEncoder
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# export PROD_DB_HOST=host
# export PROD_DB_USER=user
# export PROD_DB_PASSWORD=pass
# export PROD_DB_NAME=db
# export TEST_DB_HOST=host
# export TEST_DB_USER=user
# export TEST_DB_PASSWORD=pass
# export TEST_DB_NAME=db
# export FEISHU_RECEIVE_ID=rid
# export FEISHU_APP_ID=aid
# export FEISHU_APP_SECRET=ast

# 配置日志记录
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 从环境变量中获取配置信息
prod_db_config = {
    'host': os.getenv('PROD_DB_HOST'),
    'user': os.getenv('PROD_DB_USER'),
    'password': os.getenv('PROD_DB_PASSWORD'),
    'database': os.getenv('PROD_DB_NAME')
}

test_db_config = {
    'host': os.getenv('TEST_DB_HOST'),
    'user': os.getenv('TEST_DB_USER'),
    'password': os.getenv('TEST_DB_PASSWORD'),
    'database': os.getenv('TEST_DB_NAME')
}

receive_id = os.getenv('FEISHU_RECEIVE_ID')
app_id = os.getenv('FEISHU_APP_ID')
app_secret = os.getenv('FEISHU_APP_SECRET')

# 获取 tenant_access_token
def get_token(app_id, app_secret):
    logging.info("获取 tenant_access_token 开始")
    url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
    data = json.dumps({
        "app_id": app_id,
        "app_secret": app_secret
    })
    headers = {
        "Content-Type": "application/json; charset=utf-8"
    }
    response = requests.post(url, headers=headers, data=data)
    response_data = response.json()
    if 'tenant_access_token' in response_data:
        logging.info("获取 tenant_access_token 成功")
        return response_data['tenant_access_token']
    else:
        logging.error("获取 tenant_access_token 失败: {}".format(response_data))
        raise ValueError("Failed to get tenant access token: {}".format(response_data))

# 获取 chat_id
def get_chat_id(app_id, app_secret):
    logging.info("获取 chat_id 开始")
    url = "https://open.feishu.cn/open-apis/im/v1/chats?page_size=20"
    tenant_access_token = get_token(app_id, app_secret)
    headers = {
        'Authorization': 'Bearer ' + tenant_access_token
    }
    response = requests.get(url, headers=headers)
    response_data = response.json()
    if 'data' in response_data and 'items' in response_data['data']:
        logging.info("获取 chat_id 成功")
        return response_data['data']['items'][0]['chat_id']
    else:
        logging.error("获取 chat_id 失败: {}".format(response_data))
        raise ValueError("Failed to get chat id: {}".format(response_data))

# 获取文件二进制路径
def get_filepath(file_type, file_name, file_path, app_id, app_secret, type_file):
    logging.info("获取文件二进制路径 开始")
    url = "https://open.feishu.cn/open-apis/im/v1/files"
    with open(file_path, 'rb') as file:
        form = {
            'file_type': file_type,
            'file_name': file_name,
            'file': (file_name, file, type_file)
        }
        multi_form = MultipartEncoder(form)
        tenant_access_token = get_token(app_id, app_secret)
        headers = {
            'Authorization': 'Bearer ' + tenant_access_token,
            'Content-Type': multi_form.content_type
        }
        response = requests.post(url, headers=headers, data=multi_form)
        response_data = response.json()
        if 'data' in response_data:
            logging.info("获取文件二进制路径 成功")
            return response_data['data']['file_key']
        else:
            logging.error("获取文件二进制路径 失败: {}".format(response_data))
            raise ValueError("Failed to get file key: {}".format(response_data))

# 发送文件
def send_file_to_feishu(file_type, file_name, file_path, app_id, app_secret, type_file):
    logging.info("发送文件到飞书 开始")
    url = "https://open.feishu.cn/open-apis/im/v1/messages"
    params = {"receive_id_type": "chat_id"}
    file_key = get_filepath(file_type, file_name, file_path, app_id, app_secret, type_file)
    req = {
        "receive_id": receive_id,
        "msg_type": "file",
        "content": json.dumps({"file_key": file_key})
    }
    payload = json.dumps(req)
    tenant_access_token = get_token(app_id, app_secret)
    headers = {
        'Authorization': 'Bearer ' + tenant_access_token,
        'Content-Type': 'application/json'
    }
    response = requests.post(url, params=params, headers=headers, data=payload)
    logging.info("发送文件到飞书 结束")
    return response

# 获取数据库连接
def get_db_connection(config):
    logging.info(f"连接到数据库: {config['database']} 开始")
    connection = pymysql.connect(
        host=config['host'],
        user=config['user'],
        password=config['password'],
        database=config['database']
    )
    logging.info(f"连接到数据库: {config['database']} 成功")
    return connection

# 获取表结构
def get_table_structure(connection, table_name):
    logging.info(f"获取表结构: {table_name} 开始")
    with connection.cursor() as cursor:
        cursor.execute(f"DESCRIBE {table_name}")
        structure = cursor.fetchall()
    logging.info(f"获取表结构: {table_name} 成功")
    return structure

# 获取表索引
def get_table_indexes(connection, table_name):
    logging.info(f"获取表索引: {table_name} 开始")
    with connection.cursor() as cursor:
        cursor.execute(f"SHOW CREATE TABLE {table_name}")
        result = cursor.fetchone()
        create_table_sql = result[1]
    
    # 提取索引定义部分
    indexes = []
    for line in create_table_sql.split('\n'):
        line = line.strip()
        if line.startswith('KEY') or line.startswith('UNIQUE KEY') or line.startswith('PRIMARY KEY'):
            # 去除索引定义末尾的逗号
            index = line.rstrip(',')
            indexes.append(index)
    
    logging.info(f"获取表索引: {table_name} 成功")
    return indexes

# 获取所有表名
def get_all_tables(connection):
    logging.info("获取所有表名 开始")
    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES")
        tables = [table[0] for table in cursor.fetchall()]
    logging.info("获取所有表名 成功")
    return tables

# 比较表结构和索引
def compare_table_structures_and_indexes(prod_structure, test_structure, prod_indexes, test_indexes):
    differences = []
    prod_columns = {col[0]: col for col in prod_structure}
    test_columns = {col[0]: col for col in test_structure}
    
    for column in prod_columns:
        if column not in test_columns:
            differences.append(f"列 '{column}' 在测试数据库中缺失")
        else:
            prod_col = prod_columns[column]
            test_col = test_columns[column]
            
            # 忽略某些元数据的差异
            if (prod_col[0], prod_col[1], prod_col[2]) == (test_col[0], test_col[1], test_col[2]):
                continue
            
            differences.append(f"列 '{column}' 不一致;\n生产环境: {prod_col}\n测试环境: {test_col}")
    
    for column in test_columns:
        if column not in prod_columns:
            differences.append(f"列 '{column}' 在生产数据库中缺失")

    prod_index_set = set(prod_indexes)
    test_index_set = set(test_indexes)

    for index in prod_index_set:
        if index not in test_index_set:
            differences.append(f"索引 '{index}' 在测试数据库中缺失")

    for index in test_index_set:
        if index not in prod_index_set:
            differences.append(f"索引 '{index}' 在生产数据库中缺失")

    return differences


# 自动调整列宽和行高
def adjust_dimensions(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for row in ws.iter_rows():
        max_height = max(cell.value.count('\n') + 1 if cell.value else 1 for cell in row)
        ws.row_dimensions[row[0].row].height = max_height * 15

# 主函数
def main():
    logging.info("主函数开始")
    prod_conn = get_db_connection(prod_db_config)
    test_conn = get_db_connection(test_db_config)

    prod_tables = set(get_all_tables(prod_conn))
    test_tables = set(get_all_tables(test_conn))

    all_tables = prod_tables.union(test_tables)

    # 获取当前时间并格式化
    current_time = datetime.now().strftime("%Y%m%d")
    file_name = f"数据库结构差异报告-{prod_db_config['database']}-{current_time}.csv"

    # 保存报告到本地文件
    logging.info("保存报告到本地文件 开始")
    file_path = file_name
    with open(file_path, "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["表名", "是否存在差异", "差异的内容"])

        for table in all_tables:
            if table not in prod_tables:
                writer.writerow([table, "有差异", "表在生产数据库中缺失"])
            elif table not in test_tables:
                writer.writerow([table, "有差异", "表在测试数据库中缺失"])
            else:
                prod_structure = get_table_structure(prod_conn, table)
                test_structure = get_table_structure(test_conn, table)
                prod_indexes = get_table_indexes(prod_conn, table)
                test_indexes = get_table_indexes(test_conn, table)
                table_differences = compare_table_structures_and_indexes(prod_structure, test_structure, prod_indexes, test_indexes)
                if table_differences:
                    differences = "\n".join([f"{i+1}. {diff}" for i, diff in enumerate(table_differences)])
                    writer.writerow([table, "有差异", differences])
                else:
                    writer.writerow([table, "无差异", "无"])
    logging.info("保存报告到本地文件 结束")

    # 将 CSV 文件转换为 Excel 文件并调整列宽和行高
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(file_path, 'r', encoding='utf-8') as f:
        for row in csv.reader(f):
            ws.append(row)

    # 手动设置每列的宽度像素*8（20*8=160像素）
    ws.column_dimensions[get_column_letter(1)].width = 40
    ws.column_dimensions[get_column_letter(2)].width = 15
    ws.column_dimensions[get_column_letter(3)].width = 125

    # 设置表头字体加粗加大
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)

    # 设置单元格对齐方式，并调整行高
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            if cell.column == 2:  # "是否存在差异"列
                if cell.value == "有差异":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif cell.value == "无差异":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        max_height = max(cell.value.count('\n') + 1 if cell.value else 1 for cell in row)
        ws.row_dimensions[row[0].row].height = max_height * 15

    excel_file_name = f"数据库结构差异报告-{prod_db_config['database']}-{current_time}.xlsx"
    wb.save(excel_file_name)

    # 发送文件到飞书
    send_file_to_feishu('xlsx', excel_file_name, excel_file_name, app_id, app_secret, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # 关闭连接
    logging.info("关闭数据库连接")
    prod_conn.close()
    test_conn.close()

if __name__ == "__main__":
    main()
